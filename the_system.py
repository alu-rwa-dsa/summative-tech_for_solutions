# import openpyxl and tkinter modules
# importing Mapping to avoid any errors
from collections.abc import Mapping
from openpyxl import *
from tkinter import *
from tkinter import messagebox
import queue
import random
admin_passcode = 9043
# create the document variable
applications_document = load_workbook('NSO_applications.xlsx')
members_document = load_workbook('NSO_members.xlsx')
#an excel function to handle the excel document
applications = [{'form number': 453, 'first name': 'Alec', 'Middle name': '', 'Last name': 'Dumbuya', 'Gender': 'male',
                 'Email ID': 'gd.gmail.com', 'Phone number': '0785963843', 'Address': 'Kicukiro, Kigali', 'Occupation': 'student'},
                {'form number': 76, 'first name': 'Gideon', 'Middle name': 'Agnes', 'Last name': 'Luthor', 'Gender': 'female',
                 'Email ID': 'Gideondf3@gmail.com', 'Phone number': '0758695484', 'Address': 'Gicumbi, Rwanda', 'Occupation': 'farmer'}]
members = [{'form number': 22, 'first name': 'Queen', 'Middle name': 'Keza', 'Last name': 'Umunyana', 'Gender': 'female',
           'Email ID': 'queenumunyana29@gmail.com', 'Phone number': '0784828258', 'Address': 'Kimironko, Kigali', 'Occupation': 'student'},
           {'form number': 30, 'first name': 'Becks', 'Middle name': 'Furaha', 'Last name': 'Nishinda', 'Gender': 'non binary',
            'Email ID': 'becksfinda@gmail.com', 'Phone number': '0784953735', 'Address': 'Remera, Kigali', 'Occupation': 'photographer'}]

rejected_applications = []
sheet = applications_document.active
sheet_members = members_document.active
details = {'form number': '', 'first name': '', 'Middle name': '', 'Last name': '', 'Gender': '',
           'Email ID': '', 'Phone number': '', 'Address': '', 'Occupation': ''}
def fetch_app_data():
    row_ch = 1
    row_status = 'EMPTY DATA'
    while row_status == 'EMPTY DATA':
        try:
            row_ch = row_ch + 1
            row_view = 'NOT EMPTY'
            col_ch = 0
            cell = 1
            ch_d = []
            while cell <= len(details):
                col_ch = col_ch + 1
                value = sheet.cell(row=row_ch, column=col_ch).value
                ch_d.append(value)
                if cell == 9:
                    ch_cell = 0
                    for val in range(1, len(ch_d) + 1):
                        if ch_d[val - 1] is None:
                            ch_cell = ch_cell + 1
                            if val == 9:
                                if ch_cell == 9:
                                    row_view = 'EMPTY'
                                    row_status = 'DATA EXISTS'
                    if row_view == 'NOT EMPTY':
                        cl = 0
                        for i in details:
                            cl = cl + 1
                            value_s = sheet.cell(row=row_ch, column=cl).value
                            if value_s is None:
                                details[i] = ''
                            else:
                                details[i] = value_s
                        applications.append(details)
                cell = cell + 1
        except ValueError:
            print("Error")


def fetch_member_data():
    row_ch = 1
    row_status = 'EMPTY DATA'
    while row_status == 'EMPTY DATA':
        try:
            row_ch = row_ch + 1
            row_view = 'NOT EMPTY'
            col_ch = 0
            cell = 1
            ch_d = []
            while cell <= len(details):
                col_ch = col_ch + 1
                value = sheet_members.cell(row=row_ch, column=col_ch).value
                ch_d.append(value)
                if cell == 9:
                    ch_cell = 0
                    for val in range(1, len(ch_d) + 1):
                        if ch_d[val - 1] is None:
                            ch_cell = ch_cell + 1
                            if val == 9:
                                if ch_cell == 9:
                                    row_view = 'EMPTY'
                                    row_status = 'DATA EXISTS'
                    if row_view == 'NOT EMPTY':
                        cl = 0
                        for i in details:
                            cl = cl + 1
                            value_s = sheet_members.cell(row=row_ch, column=cl).value
                            if value_s is None:
                                details[i] = ''
                            else:
                                details[i] = value_s
                        print(details, "\n")
                        members.append(details)
                cell = cell + 1
        except ValueError:
            print("Error")
